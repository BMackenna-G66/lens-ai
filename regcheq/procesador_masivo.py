#!/usr/bin/env python3
"""
PROCESADOR MASIVO — Regcheq + Motor de Decisión
================================================
Lee un Excel con personas (DNI/RUT), llama a la API de Regcheq para
cada una, aplica el motor de decisión local y genera un Excel con
todos los resultados enriquecidos.

Modos de operación:
  1. Masivo simple  : el Excel solo tiene columnas de identificación.
  2. Base de casos  : el Excel incluye columnas crimen_N con registros judiciales.

Uso:
  python procesador_masivo.py --input masivo.xlsx --output resultado.xlsx
  python procesador_masivo.py --input base_casos.xlsx --output resultado.xlsx --modo casos
  python procesador_masivo.py --input masivo.xlsx --output resultado.xlsx --crear-fichas
  python procesador_masivo.py --input masivo.xlsx --output resultado.xlsx --dry-run
"""

import argparse
import json
import os
import sys
import time
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import regcheq_client as api
from motor_decision import evaluar_crimenes

# ─── COLORES (Excel) ─────────────────────────────────────────────────────────
COLOR_ROJO    = "FFD9D9"
COLOR_AMARILLO= "FFF2CC"
COLOR_VERDE   = "E2EFDA"
COLOR_GRIS    = "F2F2F2"
COLOR_HEADER  = "2E4057"
COLOR_FB      = "C00000"   # rojo oscuro
COLOR_UCR     = "FF6600"   # naranja
COLOR_LIB     = "375623"   # verde oscuro

RIESGO_FILL = {
    "high":   PatternFill("solid", fgColor=COLOR_ROJO),
    "High Risk": PatternFill("solid", fgColor=COLOR_ROJO),
    "medium": PatternFill("solid", fgColor=COLOR_AMARILLO),
    "low":    PatternFill("solid", fgColor=COLOR_VERDE),
}

DECISION_FILL = {
    "FORZAR_BLOQUEO":          PatternFill("solid", fgColor="FFD9D9"),
    "UNDER_COMPLIANCE_REVIEW": PatternFill("solid", fgColor="FFF2CC"),
    "Liberar":                 PatternFill("solid", fgColor="E2EFDA"),
}

DECISION_FONT = {
    "FORZAR_BLOQUEO":          Font(bold=True, color=COLOR_FB),
    "UNDER_COMPLIANCE_REVIEW": Font(bold=True, color=COLOR_UCR),
    "Liberar":                 Font(bold=True, color=COLOR_LIB),
}

# ─── MAPEO DE LISTAS REGCHEQ → COLUMNAS EXCEL ────────────────────────────────
MAPA_LISTAS = {
    "pepChile":       "Coincidencia_PEP Chile",
    "funcPublicChile":           "Coincidencia_Funcionarios Públicos Chile",
    "pdiResult":                 "Coincidencia_PDI",
    "rtpResult":                 "Coincidencia_RTP",
    "gafiResult":                "Coincidencia_Países Sancionados (GAFI)",
    "internList":                "Coincidencia_Lista de interés",
    "regcheqList":               "Coincidencia_Lista Regcheq",
    "bicResult":                 "Coincidencia_BIC",
    "keywordsResult":            "Coincidencia_Palabras Clave",
    "internationalOrganizations":"Coincidencia_Organismos internacionales",
    "ofacAddressResult":         "Coincidencia_OFAC Domicilio",
    "screeningGlobal":           "Coincidencia_Screening Global",
    "secondCriminalCasesChile":  "Coincidencia_Causas penales Chile",
    "riskComments":              "Coincidencia_Comentarios de Riesgo",
}

COLUMNAS_SALIDA = [
    "DNI",
    "Tipo de persona",
    "Nombre",
    "Apellido paterno",
    "Razón Social",
    "Riesgo calculado listas",
    "Riesgo sobreescrito",
    "Riesgo final Ficha",
    "listas_total_coincidencias",
    "Coincidencia_PEP Chile",
    "Coincidencia_Funcionarios Públicos Chile",
    "Coincidencia_Causas penales Chile",
    "Coincidencia_PDI",
    "Coincidencia_Países Sancionados (GAFI)",
    "Coincidencia_Organismos internacionales",
    "Coincidencia_OFAC Domicilio",
    "Coincidencia_Screening Global",
    "Coincidencia_RTP",
    "Coincidencia_Palabras Clave",
    "Coincidencia_Comentarios de Riesgo",
    "Coincidencia_Lista de interés",
    "Coincidencia_Lista Regcheq",
    "Coincidencia_BIC",
    "PEP_nivel",
    "causas_penales_imputado",
    "causas_penales_data",
    "regcheq_error",
    # Columnas de decisión (solo modo casos)
    "Decision",
    "Decision_Razon",
    "Precedentes_count",
    "NoPrecedentes_count",
    "Total_equivalente",
]

# Columnas que NO van a la hoja principal (se usan internamente)
COLUMNAS_INTERNAS = {"causas_penales_data"}


# ─── LECTURA DE INPUT ────────────────────────────────────────────────────────

def leer_excel_input(path: str) -> tuple[list[dict], bool]:
    """
    Lee el Excel de entrada. Detecta automáticamente si tiene columnas crimen_N.
    Retorna (lista_de_filas_como_dict, tiene_crimenes: bool).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Preferir hoja con más columnas (la de datos crudos si hay varias)
    ws_names = wb.sheetnames
    ws = wb[ws_names[0]]
    if len(ws_names) > 1:
        # Elegir la hoja con más columnas
        max_cols = ws.max_column or 0
        for name in ws_names[1:]:
            w = wb[name]
            if (w.max_column or 0) > max_cols:
                ws = w
                max_cols = w.max_column or 0

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return [], False

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    tiene_crimenes = any(h.startswith("crimen_") for h in headers)

    datos = []
    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue
        d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        datos.append(d)

    return datos, tiene_crimenes


def extraer_dni(fila: dict) -> str | None:
    for col in ("rut", "DNI", "dni", "RUT", "Rut"):
        v = fila.get(col)
        if v is not None:
            return str(v).strip()
    return None


def extraer_crimenes(fila: dict) -> list[str]:
    """Extrae los valores de columnas crimen_N (no nulos)."""
    crimenes = []
    i = 0
    while True:
        v = fila.get(f"crimen_{i}")
        if v is None and f"crimen_{i}" not in fila:
            break
        if v is not None and str(v).strip():
            crimenes.append(str(v).strip())
        i += 1
    return crimenes


# ─── CONSULTA REGCHEQ ────────────────────────────────────────────────────────

def consultar_persona(fila: dict, crear_ficha: bool) -> tuple[dict, str | None]:
    """
    Consulta Regcheq para una persona. Retorna (resultado_normalizado, error_str).
    resultado_normalizado tiene las mismas claves que COLUMNAS_SALIDA.
    """
    dni = extraer_dni(fila)
    if not dni:
        return {}, "DNI no encontrado en la fila"

    # Datos de contexto del input
    nombre   = fila.get("Nombre") or fila.get("nombre") or ""
    apellido = fila.get("Apellido") or fila.get("Apellido paterno") or fila.get("apellido") or ""
    razon    = fila.get("Razón Social") or fila.get("razon_social") or ""
    tipo_p   = str(fila.get("Tipo de persona") or "natural").lower()

    resultado = {
        "DNI":             dni,
        "Tipo de persona": tipo_p,
        "Nombre":          nombre,
        "Apellido paterno": apellido,
        "Razón Social":    razon,
    }

    try:
        if crear_ficha:
            datos_ficha = {"dni": dni, "personType": tipo_p}
            if nombre:
                datos_ficha["name"] = str(nombre).upper()
            if apellido:
                datos_ficha["fatherName"] = str(apellido).upper()
            if razon:
                datos_ficha["socialReason"] = razon
            api.crear_ficha(datos_ficha)
            time.sleep(0.3)   # pequeña pausa tras crear

        perfil = api.obtener_ficha(dni)
    except Exception as e:
        return resultado, str(e)

    # ─ Extraer campos de riesgo ─
    riesgo_calc = perfil.get("calculatedRisk") or perfil.get("riesgo_calculado") or ""
    riesgo_over = perfil.get("overwrittenRisk") or ""
    riesgo_final= perfil.get("effectiveRisk")   or perfil.get("riesgo_efectivo") or riesgo_calc
    pep_level   = perfil.get("pepLevel") or ""

    resultado["Riesgo calculado listas"] = riesgo_calc
    resultado["Riesgo sobreescrito"]     = riesgo_over
    resultado["Riesgo final Ficha"]      = riesgo_final
    resultado["PEP_nivel"]               = pep_level

    # ─ Extraer coincidencias de listas ─
    listas = perfil.get("listas") or {}
    total_coincidencias = 0
    for clave_api, col_excel in MAPA_LISTAS.items():
        entry = listas.get(clave_api) or {}
        coincide = bool(entry.get("coincidence", False))
        resultado[col_excel] = coincide
        if coincide:
            total_coincidencias += 1

    resultado["listas_total_coincidencias"] = str(total_coincidencias)

    # ─ Extraer causas penales con detalle completo ─
    causas_entry = listas.get("secondCriminalCasesChile") or {}
    if isinstance(causas_entry, dict) and causas_entry.get("coincidence"):
        raw = causas_entry.get("data") or {}
        if isinstance(raw, dict):
            casos = raw.get("additionalData") or []
            info  = raw.get("info") or {}
            resultado["causas_penales_data"]     = json.dumps(casos if isinstance(casos, list) else [], ensure_ascii=False)
            resultado["causas_penales_imputado"] = info.get("name") or ""
        else:
            resultado["causas_penales_data"]     = ""
            resultado["causas_penales_imputado"] = ""
    else:
        resultado["causas_penales_data"]     = ""
        resultado["causas_penales_imputado"] = ""

    # Nombre / apellido desde perfil si no vinieron en el input
    if not resultado["Nombre"]:
        resultado["Nombre"] = perfil.get("name") or ""
    if not resultado["Apellido paterno"]:
        resultado["Apellido paterno"] = perfil.get("fatherName") or ""
    if not resultado["Razón Social"]:
        resultado["Razón Social"] = perfil.get("socialReason") or ""
    if not resultado["Tipo de persona"]:
        resultado["Tipo de persona"] = perfil.get("personType") or tipo_p

    return resultado, None


# ─── ESCRITURA DE OUTPUT ──────────────────────────────────────────────────────

def _header_style(ws, fila=1):
    fill = PatternFill("solid", fgColor=COLOR_HEADER)
    font = Font(bold=True, color="FFFFFF", size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        bottom=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
    )
    for cell in ws[fila]:
        cell.fill  = fill
        cell.font  = font
        cell.alignment = align
        cell.border = border


def _autofit(ws, max_width=40):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, max_width)


def _aplicar_colores_fila(ws, row_idx: int, fila_dict: dict, cols: list[str]):
    riesgo = str(fila_dict.get("Riesgo final Ficha") or "").lower()
    decision = fila_dict.get("Decision") or ""

    # Color de fondo por riesgo en toda la fila
    fill = RIESGO_FILL.get(riesgo) or RIESGO_FILL.get(fila_dict.get("Riesgo final Ficha"))
    if fill:
        for c_idx in range(1, len(cols) + 1):
            ws.cell(row=row_idx, column=c_idx).fill = fill

    # Celda de decisión con formato propio
    if decision and "Decision" in cols:
        col_idx = cols.index("Decision") + 1
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill  = DECISION_FILL.get(decision, PatternFill())
        cell.font  = DECISION_FONT.get(decision, Font())
        cell.alignment = Alignment(horizontal="center")

    # Columnas de coincidencia booleana → TRUE en rojo
    for col_name in cols:
        if col_name.startswith("Coincidencia_"):
            c_idx = cols.index(col_name) + 1
            cell  = ws.cell(row=row_idx, column=c_idx)
            if cell.value is True or str(cell.value).upper() == "TRUE":
                cell.fill = PatternFill("solid", fgColor=COLOR_ROJO)
                cell.font = Font(bold=True)


def escribir_excel_output(path: str, resultados: list[dict], tiene_crimenes: bool):
    wb = openpyxl.Workbook()

    # ─── Hoja principal ───
    ws = wb.active
    ws.title = "Resultados Regcheq"
    ws.freeze_panes = "A2"

    cols_usar = [c for c in COLUMNAS_SALIDA if (
        c not in ("Decision", "Decision_Razon", "Precedentes_count",
                  "NoPrecedentes_count", "Total_equivalente")
        or tiene_crimenes
    ) and c not in COLUMNAS_INTERNAS]

    ws.append(cols_usar)
    _header_style(ws)

    for row_idx, fila in enumerate(resultados, start=2):
        row_vals = [fila.get(c, "") for c in cols_usar]
        ws.append(row_vals)
        _aplicar_colores_fila(ws, row_idx, fila, cols_usar)

    _autofit(ws)
    ws.row_dimensions[1].height = 30

    # ─── Hoja de coincidencias (solo filas con al menos 1 alerta) ───
    cols_coinc = ["DNI", "Nombre", "Apellido paterno", "Razón Social",
                  "Riesgo final Ficha", "causas_penales_imputado",
                  "Coincidencia_Causas penales Chile",
                  "Coincidencia_PEP Chile",
                  "Coincidencia_Funcionarios Públicos Chile",
                  "Coincidencia_PDI",
                  "Coincidencia_Países Sancionados (GAFI)",
                  "Coincidencia_Organismos internacionales",
                  "Coincidencia_OFAC Domicilio",
                  "Coincidencia_Screening Global",
                  "Coincidencia_RTP",
                  "Coincidencia_BIC",
                  "Coincidencia_Palabras Clave",
                  "Coincidencia_Comentarios de Riesgo",
                  "Coincidencia_Lista de interés",
                  "Coincidencia_Lista Regcheq"]
    registros_con_alerta = [
        r for r in resultados
        if any(r.get(c) is True for c in cols_coinc if c.startswith("Coincidencia_"))
    ]
    ws_c = wb.create_sheet("Coincidencias")
    ws_c.freeze_panes = "A2"
    cabecera_c = [c.replace("causas_penales_imputado", "Nombre imputado (API)") for c in cols_coinc]
    ws_c.append(cabecera_c)
    _header_style(ws_c)
    FILL_CP_HEADER = PatternFill("solid", fgColor="7B0000")   # rojo muy oscuro
    FONT_CP_HEADER = Font(bold=True, color="FFFFFF", size=10)
    # Destacar columna de causas penales en header
    cp_col_idx = cols_coinc.index("Coincidencia_Causas penales Chile") + 1
    ws_c.cell(row=1, column=cp_col_idx).fill = FILL_CP_HEADER
    ws_c.cell(row=1, column=cp_col_idx).font = FONT_CP_HEADER

    for row_idx, fila in enumerate(registros_con_alerta, start=2):
        vals = [fila.get(c, "") for c in cols_coinc]
        ws_c.append(vals)
        _aplicar_colores_fila(ws_c, row_idx, fila, cols_coinc)
        # Celda de causas penales: rojo intenso si es True
        cell_cp = ws_c.cell(row=row_idx, column=cp_col_idx)
        if cell_cp.value is True:
            cell_cp.fill = PatternFill("solid", fgColor="C00000")
            cell_cp.font = Font(bold=True, color="FFFFFF")
    _autofit(ws_c)
    ws_c.row_dimensions[1].height = 30

    # ─── Hoja causas penales detallada (1 fila por delito) ───
    registros_con_causas = [r for r in resultados if r.get("causas_penales_data")]
    if registros_con_causas:
        ws_cp = wb.create_sheet("Causas Penales Chile")
        ws_cp.freeze_panes = "A2"
        hdrs_cp = ["DNI", "Imputado (API)", "Nombre Ficha", "Riesgo Ficha",
                   "Delito", "Estado", "Fecha", "Riesgo Delito", "RIT", "RUC", "Tribunal"]
        ws_cp.append(hdrs_cp)
        # Header estilo especial para causas
        fill_hdr = PatternFill("solid", fgColor="7B0000")
        font_hdr = Font(bold=True, color="FFFFFF", size=10)
        align_hdr = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws_cp[1]:
            cell.fill, cell.font, cell.alignment = fill_hdr, font_hdr, align_hdr
        ws_cp.row_dimensions[1].height = 30

        FILL_H = PatternFill("solid", fgColor="FFD9D9")
        FILL_M = PatternFill("solid", fgColor="FFF2CC")
        FILL_L = PatternFill("solid", fgColor="E2EFDA")

        for fila in registros_con_causas:
            try:
                casos = json.loads(fila["causas_penales_data"])
            except Exception:
                continue
            # Deduplicar por RUC (conservar el más reciente)
            ruc_map = {}
            for caso in casos:
                ruc = caso.get("ruc", "")
                fecha_str = caso.get("fecha", "")
                try:
                    from datetime import datetime as dt
                    fecha_dt = dt.strptime(fecha_str, "%d/%m/%Y")
                except Exception:
                    fecha_dt = None
                if ruc not in ruc_map or (fecha_dt and ruc_map[ruc][1] and fecha_dt > ruc_map[ruc][1]):
                    ruc_map[ruc] = (caso, fecha_dt)
            casos_dedup = [v[0] for v in ruc_map.values()]

            for caso in casos_dedup:
                row_vals = [
                    fila.get("DNI", ""),
                    fila.get("causas_penales_imputado", ""),
                    f"{fila.get('Nombre','')} {fila.get('Apellido paterno','')}".strip(),
                    fila.get("Riesgo final Ficha", ""),
                    caso.get("crimen", ""),
                    caso.get("estado", ""),
                    caso.get("fecha", ""),
                    caso.get("riesgo", "").upper() if caso.get("riesgo") else "",
                    caso.get("rit", ""),
                    caso.get("ruc", ""),
                    caso.get("tribunal", ""),
                ]
                ws_cp.append(row_vals)
                ri = ws_cp.max_row
                riesgo_d = str(caso.get("riesgo", "")).lower()
                fill_d = FILL_H if riesgo_d == "high" else FILL_M if riesgo_d == "medium" else FILL_L if riesgo_d == "low" else None
                if fill_d:
                    for ci in range(1, len(hdrs_cp) + 1):
                        ws_cp.cell(row=ri, column=ci).fill = fill_d
        _autofit(ws_cp)

    # ─── Hoja de resumen ───
    ws2 = wb.create_sheet("Resumen")
    total     = len(resultados)
    errores   = sum(1 for r in resultados if r.get("regcheq_error"))
    high_risk = sum(1 for r in resultados if str(r.get("Riesgo final Ficha") or "").lower() == "high risk")
    pep       = sum(1 for r in resultados if r.get("Coincidencia_PEP Chile"))
    fb        = sum(1 for r in resultados if r.get("Decision") == "FORZAR_BLOQUEO")
    ucr       = sum(1 for r in resultados if r.get("Decision") == "UNDER_COMPLIANCE_REVIEW")
    liberar   = sum(1 for r in resultados if r.get("Decision") == "Liberar")

    ws2_data = [
        ("Generado",          datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Total personas",    total),
        ("Errores API",       errores),
        ("High Risk",         high_risk),
        ("PEP detectado",     pep),
        ("",                  ""),
        ("FORZAR_BLOQUEO",    fb),
        ("UNDER_COMPLIANCE_REVIEW", ucr),
        ("Liberar",           liberar),
    ]
    for r in ws2_data:
        ws2.append(r)

    if tiene_crimenes and total > 0:
        ws2.append(("",))
        ws2.append(("Nota", "Decisiones calculadas con Catálogo de Delitos + Tabla de Decisión local"))

    _autofit(ws2, max_width=40)

    wb.save(path)
    print(f"  Guardado: {path}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Procesamiento masivo de personas contra Regcheq API + Motor de Decisión",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  # Procesar listado simple de DNIs
  python procesador_masivo.py --input masivo.xlsx --output resultado.xlsx

  # Procesar base de casos con crímenes (aplica también motor de decisión)
  python procesador_masivo.py --input base_casos.xlsx --output resultado.xlsx

  # Crear fichas antes de consultar (útil si los DNIs son nuevos en Regcheq)
  python procesador_masivo.py --input masivo.xlsx --output resultado.xlsx --crear-fichas

  # Solo mostrar lo que haría, sin llamar a la API
  python procesador_masivo.py --input masivo.xlsx --output resultado.xlsx --dry-run

  # Limitar cantidad de filas para prueba
  python procesador_masivo.py --input masivo.xlsx --output resultado.xlsx --limite 10
        """
    )
    parser.add_argument("--input",  "-i", required=True, help="Excel de entrada")
    parser.add_argument("--output", "-o", required=True, help="Excel de salida")
    parser.add_argument("--crear-fichas", action="store_true",
                        help="Crea/actualiza ficha en Regcheq antes de consultar")
    parser.add_argument("--dry-run", action="store_true",
                        help="Solo lee el input, no llama a la API")
    parser.add_argument("--limite", type=int, default=0,
                        help="Máximo de filas a procesar (0 = sin límite)")
    parser.add_argument("--delay", type=float, default=0.5,
                        help="Segundos de espera entre llamadas API (default: 0.5)")
    args = parser.parse_args()

    # ─ Validar paths ─
    if not os.path.exists(args.input):
        print(f"ERROR: No se encontró el archivo de entrada: {args.input}")
        sys.exit(1)

    print(f"\nLeyendo: {args.input}")
    filas, tiene_crimenes = leer_excel_input(args.input)
    print(f"  {len(filas)} personas encontradas")
    print(f"  Modo: {'Base de Casos (con crímenes)' if tiene_crimenes else 'Masivo Simple'}")

    if args.limite:
        filas = filas[:args.limite]
        print(f"  Limitado a: {args.limite} filas")

    if args.dry_run:
        print("\n[DRY RUN] No se llamará a la API.")
        for i, f in enumerate(filas[:5], 1):
            print(f"  {i}. DNI={extraer_dni(f)}  nombre={f.get('Nombre','')} {f.get('Apellido','')}")
        return

    # ─ Procesar ─
    resultados = []
    errores_count = 0

    print(f"\nProcesando {'con creación de fichas' if args.crear_fichas else 'solo consulta'}...\n")

    for i, fila in enumerate(filas, 1):
        dni = extraer_dni(fila)
        nombre_display = f"{fila.get('Nombre','')} {fila.get('Apellido', fila.get('Apellido paterno',''))}".strip()
        print(f"  [{i:>4}/{len(filas)}] DNI: {dni:<15} {nombre_display:<30}", end="", flush=True)

        resultado, error = consultar_persona(fila, args.crear_fichas)

        if error:
            resultado["regcheq_error"] = error
            resultado.setdefault("DNI", dni)
            errores_count += 1
            print(f"  ERROR: {error[:60]}")
        else:
            resultado["regcheq_error"] = ""

            # Aplicar motor de decisión si hay crímenes
            if tiene_crimenes:
                crimenes = extraer_crimenes(fila)
                if crimenes:
                    eval_result = evaluar_crimenes(crimenes)
                    resultado["Decision"]            = eval_result["decision"]
                    resultado["Decision_Razon"]      = eval_result["razon"]
                    resultado["Precedentes_count"]   = eval_result["precedentes_count"]
                    resultado["NoPrecedentes_count"]= eval_result["noprecedentes_count"]
                    resultado["Total_equivalente"]   = eval_result["total_equivalente"]
                else:
                    resultado["Decision"] = "Liberar"
                    resultado["Decision_Razon"] = "Sin crímenes asociados"
                    resultado["Precedentes_count"]   = 0
                    resultado["NoPrecedentes_count"] = 0
                    resultado["Total_equivalente"]   = 0.0

            riesgo = resultado.get("Riesgo final Ficha") or "—"
            decision = resultado.get("Decision") or ""
            decision_str = f" | {decision}" if decision else ""
            print(f"  OK  riesgo={riesgo:<10}{decision_str}")

        resultados.append(resultado)

        # Pausa entre requests
        if i < len(filas):
            time.sleep(args.delay)

    # ─ Output ─
    print(f"\nEscribiendo resultados ({len(resultados)} filas, {errores_count} errores)...")
    escribir_excel_output(args.output, resultados, tiene_crimenes)

    print(f"\n{'─'*50}")
    print(f"Total procesado : {len(resultados)}")
    print(f"Errores API     : {errores_count}")
    ok = [r for r in resultados if not r.get("regcheq_error")]
    if ok:
        high = sum(1 for r in ok if str(r.get("Riesgo final Ficha","")).lower() == "high risk")
        pep  = sum(1 for r in ok if r.get("Coincidencia_PEP Chile"))
        print(f"High Risk       : {high}")
        print(f"PEP Chile       : {pep}")
        if tiene_crimenes:
            fb  = sum(1 for r in ok if r.get("Decision") == "FORZAR_BLOQUEO")
            ucr = sum(1 for r in ok if r.get("Decision") == "UNDER_COMPLIANCE_REVIEW")
            lib = sum(1 for r in ok if r.get("Decision") == "Liberar")
            print(f"FORZAR_BLOQUEO  : {fb}")
            print(f"UCR             : {ucr}")
            print(f"Liberar         : {lib}")
    print(f"{'─'*50}\n")


if __name__ == "__main__":
    main()
