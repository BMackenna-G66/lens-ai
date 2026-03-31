"""
Motor de Decisión — Matriz de Delitos
======================================
Carga el Catálogo de Delitos, la Tabla de Decisión y los Parámetros
desde los archivos Excel del proyecto principal y expone la función
`evaluar_crimenes(lista_crimenes)` que retorna la decisión final:
  - "Liberar"
  - "UNDER_COMPLIANCE_REVIEW"  (UCR)
  - "FORZAR_BLOQUEO"           (FB)

Los crímenes de entrada son strings con el nombre del delito tal como
vienen de la base de casos (columnas crimen_N).
"""

import os
import openpyxl

# ─── RUTAS ────────────────────────────────────────────────────────────────────
_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_CATALOGO   = os.path.join(_BASE, "Matriz de Delitos", "Catalogo_Delitos.xlsx")
_TABLA      = os.path.join(_BASE, "Matriz de Delitos", "Tabla_Decision.xlsx")
_PARAMETROS = os.path.join(_BASE, "Matriz de Delitos", "Parametros.xlsx")


# ─── CARGA DE DATOS ──────────────────────────────────────────────────────────

def _cargar_parametros() -> dict:
    wb = openpyxl.load_workbook(_PARAMETROS, read_only=True, data_only=True)
    ws = wb.active
    params = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] and row[2] is not None:
            params[row[0]] = row[2]
    wb.close()
    return params


def _cargar_catalogo() -> dict:
    """Retorna dict: nombre_delito_upper → {tipo, riesgo_regcheq, riesgo_g66, valor}"""
    wb = openpyxl.load_workbook(_CATALOGO, read_only=True, data_only=True)
    ws = wb.active
    catalogo = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre, riesgo_rcq, riesgo_g66, valor, tipo, estado = row[:6]
        if nombre and estado and str(estado).strip().lower() == "ok":
            catalogo[str(nombre).strip().upper()] = {
                "tipo":         str(tipo or "").strip(),
                "riesgo_rcq":   str(riesgo_rcq or "").strip(),
                "riesgo_g66":   str(riesgo_g66 or "").strip(),
                "valor":        float(valor) if valor is not None else 0.0,
            }
    wb.close()
    return catalogo


def _cargar_tabla_decision() -> list:
    """Retorna lista de dicts con los campos de la Tabla_Decision."""
    wb = openpyxl.load_workbook(_TABLA, read_only=True, data_only=True)
    ws = wb.active
    tabla = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        tabla.append({
            "precedentes_count":  int(row[0]),
            "noprecedentes_count":int(row[1]),
            "pre_equivalente":    float(row[2]) if row[2] is not None else 0.0,
            "nopre_equivalente":  float(row[3]) if row[3] is not None else 0.0,
            "total_equivalente":  float(row[4]) if row[4] is not None else 0.0,
            "decision":           str(row[5]).strip() if row[5] else "",
            "razon":              str(row[6]).strip() if row[6] else "",
        })
    wb.close()
    return tabla


# Cache en memoria (se carga una sola vez por proceso)
_PARAMS  = None
_CAT     = None
_TABLA_D = None


def _init():
    global _PARAMS, _CAT, _TABLA_D
    if _PARAMS is None:
        _PARAMS  = _cargar_parametros()
        _CAT     = _cargar_catalogo()
        _TABLA_D = _cargar_tabla_decision()


# ─── LÓGICA DE EVALUACIÓN ────────────────────────────────────────────────────

def clasificar_crimen(nombre_crimen: str) -> dict | None:
    """
    Busca el crimen en el catálogo (búsqueda exacta, luego parcial).
    Retorna el registro del catálogo o None si no se encuentra.
    """
    _init()
    key = str(nombre_crimen).strip().upper()
    if key in _CAT:
        return _CAT[key]
    # Búsqueda parcial (contiene)
    for nombre_cat, datos in _CAT.items():
        if key in nombre_cat or nombre_cat in key:
            return datos
    return None


def evaluar_crimenes(lista_crimenes: list[str]) -> dict:
    """
    Dada una lista de nombres de crímenes, retorna:
    {
      "decision":           "Liberar" | "UNDER_COMPLIANCE_REVIEW" | "FORZAR_BLOQUEO",
      "razon":              str,
      "precedentes_count":  int,
      "noprecedentes_count":int,
      "total_equivalente":  float,
      "detalle":            [{nombre, tipo, riesgo_rcq, riesgo_g66, valor, encontrado}]
    }
    """
    _init()

    p  = _PARAMS
    HARD_FB_PRE   = int(p.get("HARD_FB_precedentes_count",   4))
    HARD_FB_NOPRE = int(p.get("HARD_FB_noprecedentes_count", 5))
    FB_MIN_EQ     = float(p.get("FB_min_total_equivalente",  2.0))
    UCR_MIN_EQ    = float(p.get("UCR_min_total_equivalente", 1.0))
    UCR_MAX_EQ    = float(p.get("UCR_max_total_equivalente", 3.0))
    VAL_PRE       = float(p.get("VALOR_unit_precedente",     1.0))
    VAL_NOPRE_LOW = float(p.get("VALOR_unit_nopre_low",      0.5))
    VAL_NOPRE_HIGH= float(p.get("VALOR_unit_nopre_high",     1.0))

    detalle = []
    precedentes_count   = 0
    noprecedentes_count = 0
    total_equivalente   = 0.0

    crimenes_limpios = [c for c in lista_crimenes if c and str(c).strip()]

    for nombre in crimenes_limpios:
        info = clasificar_crimen(nombre)
        if info:
            tipo = info["tipo"]
            es_precedente = "precedente" in tipo.lower()
            riesgo_g66    = info["riesgo_g66"].lower()

            if es_precedente:
                precedentes_count   += 1
                equivalente          = VAL_PRE
            else:
                noprecedentes_count += 1
                equivalente = VAL_NOPRE_HIGH if riesgo_g66 == "alto" else VAL_NOPRE_LOW

            total_equivalente += equivalente
            detalle.append({
                "nombre":    nombre,
                "tipo":      tipo,
                "riesgo_rcq": info["riesgo_rcq"],
                "riesgo_g66": info["riesgo_g66"],
                "equivalente": equivalente,
                "encontrado": True,
            })
        else:
            # Crimen no encontrado en catálogo → conservador: contar como no-precedente leve
            noprecedentes_count += 1
            total_equivalente   += VAL_NOPRE_LOW
            detalle.append({
                "nombre":    nombre,
                "tipo":      "Desconocido",
                "riesgo_rcq": "—",
                "riesgo_g66": "—",
                "equivalente": VAL_NOPRE_LOW,
                "encontrado": False,
            })

    # ─ Reglas duras ─
    if precedentes_count >= HARD_FB_PRE:
        return _resultado("FORZAR_BLOQUEO",
                          f"Hard rule: {precedentes_count} delitos precedentes >= {HARD_FB_PRE}",
                          precedentes_count, noprecedentes_count, total_equivalente, detalle)
    if noprecedentes_count >= HARD_FB_NOPRE:
        return _resultado("FORZAR_BLOQUEO",
                          f"Hard rule: {noprecedentes_count} no-precedentes >= {HARD_FB_NOPRE}",
                          precedentes_count, noprecedentes_count, total_equivalente, detalle)
    if total_equivalente >= FB_MIN_EQ:
        return _resultado("FORZAR_BLOQUEO",
                          f"Total equivalente {total_equivalente:.1f} >= {FB_MIN_EQ}",
                          precedentes_count, noprecedentes_count, total_equivalente, detalle)

    # ─ Buscar en Tabla_Decision ─
    mejor = _buscar_tabla(precedentes_count, noprecedentes_count, total_equivalente)
    if mejor:
        return _resultado(mejor["decision"], mejor["razon"],
                          precedentes_count, noprecedentes_count, total_equivalente, detalle)

    # ─ Fallback por equivalente ─
    if total_equivalente >= UCR_MIN_EQ and total_equivalente < UCR_MAX_EQ:
        return _resultado("UNDER_COMPLIANCE_REVIEW",
                          f"Equivalente {total_equivalente:.1f} en rango UCR",
                          precedentes_count, noprecedentes_count, total_equivalente, detalle)

    return _resultado("Liberar", "Sin umbral de alerta",
                      precedentes_count, noprecedentes_count, total_equivalente, detalle)


def _resultado(decision, razon, pre, nopre, total, detalle):
    return {
        "decision":            decision,
        "razon":               razon,
        "precedentes_count":   pre,
        "noprecedentes_count": nopre,
        "total_equivalente":   round(total, 2),
        "detalle":             detalle,
    }


def _buscar_tabla(pre, nopre, total) -> dict | None:
    """Busca la fila más específica en la Tabla_Decision."""
    _init()
    candidatos = [
        r for r in _TABLA_D
        if r["precedentes_count"]   == pre
        and r["noprecedentes_count"] == nopre
        and abs(r["total_equivalente"] - total) < 0.05
    ]
    return candidatos[0] if candidatos else None
